import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import plotly.graph_objects as go

import pandas as pd

import pandas as pd
import streamlit as st

url = 'https://github.com/SNgere/Allocator/raw/f1e408d2a5328114bcaabc759c995cc4c6d95562/verified.xlsx'

@st.cache
def load_data(url):
    df = pd.read_excel(url, engine='openpyxl')
    return df

df = load_data(url)
st.write(df)



# Load the workbook using openpyxl
wb = openpyxl.load_workbook(url)

# Get the active worksheet
ws = wb.active

# Initialize variables
red_count = 0
not_red_count = 0

# Get the column index of the 'Date' column in the Date frame
date_col_idx = df.columns.get_loc('Date')

# Iterate over each cell in the worksheet
for row in ws.iter_rows():
    for cell in row:
        # Check if cell has a red fill color
        if isinstance(cell.fill, PatternFill) and cell.fill.start_color.index == 'FFFF0000':
            # Check if cell value is not NaN and cell is not in the 'Date' column
            if not pd.isna(cell.value) and cell.column != date_col_idx + 1:
                red_count += 1
        else:
            # Check if cell value is not NaN and cell is not in the 'Date' column
            if not pd.isna(cell.value) and cell.column != date_col_idx + 1:
                not_red_count += 1 

# Create a pie chart showing the counts of cells with and without red fill color
fig = go.Figure(data=[go.Pie(labels=['Red', 'Not Red'], values=[red_count, not_red_count], hole=.3)])

# Set the title of the pie chart
fig.update_layout(title='Counts of Cells with and Without Red Fill Color')

# Display the pie chart using Streamlit
st.plotly_chart(fig)


#############################################################################################################################################################

import streamlit as st
import pandas as pd
import numpy as np

# Set page title
st.set_page_config(
    page_title="Scanning Department",
    page_icon=":memo:",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Define page background color and font
page_bg_color = "#F8F8FF"
font = "sans-serif"

# Define header styles
header_bg_color = "#222831"
header_text_color = "#F8F8FF"
header_font_size = "30px"
header_font_weight = "bold"
header_padding = "1rem"

# Define subheader styles
subheader_bg_color = "#393e46"
subheader_text_color = "#F8F8FF"
subheader_font_size = "20px"
subheader_font_weight = "bold"
subheader_padding = "0.5rem"

# Set page header
st.markdown(
    f"""
    <style>
        .reportview-container {{
            background-color: {page_bg_color};
            font-family: {font};
        }}
        .css-hby737 {{
            padding: 0;
        }}
        .main-header {{
            background-color: {header_bg_color};
            color: {header_text_color};
            font-size: {header_font_size};
            font-weight: {header_font_weight};
            padding: {header_padding};
            margin-bottom: 1rem;
        }}
        .main-subheader {{
            background-color: {subheader_bg_color};
            color: {subheader_text_color};
            font-size: {subheader_font_size};
            font-weight: {subheader_font_weight};
            padding: {subheader_padding};
        }}
    </style>
    
    <div class="main-header">
        <div class="css-hby737"> </div>
        <span>Scanning Department</span>
    </div>
    
    <div class="main-subheader">
        <span>Batches allocated weekly</span>
    </div>
    """,
    unsafe_allow_html=True
)

# Load the dataframe from CSV file
df2 = pd.read_csv('https://raw.githubusercontent.com/SNgere/Allocator/main/job.csv')

# Convert 'Date' column to datetime format
df2['Date'] = pd.to_datetime(df2['Date'])

# Set 'Date' column as the index
df2.set_index('Date', inplace=True)

# Filter the dataframe to only show Monday to Friday
df2_weekdays = df2.loc[df2.index.weekday < 5]

# Define a function to format the date as "Monday, DD/MM/YYYY"
def format_date(date):
    return date.strftime('%A, %d/%m/%Y')

# Apply the date format to the index column
df2_weekdays.index = df2_weekdays.index.map(format_date)

# Display the dataframe in Streamlit
st.write(df2_weekdays)

#################################################################################################################################################################

import pandas as pd
import streamlit as st

# Load the data
df = pd.read_csv("https://raw.githubusercontent.com/SNgere/Allocator/main/job.csv")

# Add a text input widget to allow the user to search
search_term = st.text_input("Search")

# Convert the search term to a numeric type
try:
    search_num = int(search_term)
except ValueError:
    try:
        search_num = float(search_term)
    except ValueError:
        search_num = None

# Check if the search term matches any column
if search_num is not None:
    match_found = False
    for col in df.columns:
        if df[col].dtype in ['int64', 'float64'] and col != 'date':
            if search_num in df[col].values:
                st.write(f"{search_num} was assigned to '{col}'.")
                match_found = True
    if not match_found:
        st.write(f"No match found for {search_num}.")
else:
    st.write("Enter a valid number to search.")


